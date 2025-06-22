from gurobipy import Model, GRB
import sys
import pandas as pd
import os
import sys
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime 
from matplotlib import pyplot as plt
import fileinput

def read_file(file_path):
    s = ""
    for line in fileinput.input(files=file_path):
        s += line
    return s.splitlines()

W, H = 0, 0
global result_dict
result_dict = {}
def input_data(file_path):
    global W, H
    data = {}
    f = open(file_path,'r')
    
    data['size_item'] = []
    data['size_bin'] = []
    n = int(f.readline())
    W, H = map(int, f.readline().split())
    data['size_bin'].append([W, H])
    for i in range(1,n+1):
        line = f.readline().split()
        data['size_item'].append([int(line[0]),int(line[1])])
    
    return n,data,W,H


def display_solution(strip, rectangles, pos_circuits, rotation):
    # define Matplotlib figure and axis
    fig, ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i],
                                 rectangles[i][0] if not rotation[i] else rectangles[i][1],
                                 rectangles[i][1] if not rotation[i] else rectangles[i][0],
                                 edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    # display plot
    plt.show()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path = 'out/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)
    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")
def main_solver(file_path, time_limit):
    n,data,W,H = input_data(file_path)
    k = n
    print('Number of items:',n)
    print('Size of bin:',W,H)
     
    global result_dict 
    result_dict = {
        "Type": "Gurobi",
        'Problem': file_path.split("/")[-1],
        "Number of items": n,
        "Width": W,
        "Height": H,
        "Number of bins": 0,
        "Time": 0,
        "Result": "SAT",
        }
    # Create Gurobi model
    model = Model("BinPacking")
    
    # Create variables
    M = 1000000

    x = {} # x[(i,m)] = 1 iff item i is packed in car m else 0
    # Ro represent for R in the presentation file/ pdf model file
    Ro = {} # if Ro = 1 then rotation = 90 degree, else 0
    l = {} # left coordination of item
    r = {} # right coordination of item
    t = {} # top coordination of item
    b = {} # bottom coodination of item
    
    for i in range(n):
        Ro[i] = model.addVar(vtype=GRB.BINARY, name=f'Ro[{i}]')

        # coordinate
        l[i] = model.addVar(lb=0, ub=W, vtype=GRB.CONTINUOUS, name=f'l[{i}]')
        r[i] = model.addVar(lb=0, ub=W, vtype=GRB.CONTINUOUS, name=f'r[{i}]')
        t[i] = model.addVar(lb=0, ub=H, vtype=GRB.CONTINUOUS, name=f't[{i}]')
        b[i] = model.addVar(lb=0, ub=H, vtype=GRB.CONTINUOUS, name=f'b[{i}]')        

        model.addConstr(r[i] == (1-Ro[i]) * data['size_item'][i][0] + Ro[i] * data['size_item'][i][1] + l[i])
        model.addConstr(t[i] == (1-Ro[i]) * data['size_item'][i][1] + Ro[i] * data['size_item'][i][0] + b[i])

        for m in range(k):
            x[(i,m)] = model.addVar(vtype=GRB.BINARY, name=f'x[{i}][{m}]')

            # item i must not exceed area of car
            model.addConstr(r[i] <= (1-x[(i,m)]) * M + W)
            model.addConstr(l[i] <= (1-x[(i,m)]) * M + W)
            model.addConstr(t[i] <= (1-x[(i,m)]) * M + H)
            model.addConstr(b[i] <= (1-x[(i,m)]) * M + H)    

    for i in range(n):
        model.addConstr(sum(x[(i,m)] for m in range(k)) == 1)

    # if 2 items is packed in the same car, they must be not overlaped
    for i in range(n - 1):
        for j in range(i + 1, n):
            for m in range(k):
                e = model.addVar(vtype=GRB.BINARY, name=f'e[{i}][{j}]')
                model.addConstr(e >= x[i,m] + x[j,m] - 1)
                model.addConstr(e <= x[i,m])
                model.addConstr(e <= x[j,m])

                # Binary variables for each constraint
                c1 = model.addVar(vtype=GRB.BINARY, name=f'c1[{i}][{j}]')
                c2 = model.addVar(vtype=GRB.BINARY, name=f'c2[{i}][{j}]')
                c3 = model.addVar(vtype=GRB.BINARY, name=f'c3[{i}][{j}]')
                c4 = model.addVar(vtype=GRB.BINARY, name=f'c4[{i}][{j}]')
                
                # Constraints that the binary variables must satisfy
                model.addConstr(r[i] <= l[j] + M * (1 - c1))
                model.addConstr(r[j] <= l[i] + M * (1 - c2))
                model.addConstr(t[i] <= b[j] + M * (1 - c3))
                model.addConstr(t[j] <= b[i] + M * (1 - c4))

                model.addConstr(c1 + c2 + c3 + c4 + (1-e)*M >= 1)
                model.addConstr(c1 + c2 + c3 + c4 <= e*M)

    # find cars be used
    z = {} # z[m] = 1 iff car m be used
    for m in range(k):
        z[m] = model.addVar(vtype=GRB.BINARY, name=f'z[{m}]')
        # if sum(x[i][m]) >= 1 then car m be used => z[m] = 1
        # else, z[m] = 0

        q = model.addVar(lb=0, ub=n, vtype=GRB.INTEGER, name=f'q[{m}]')
        model.addConstr(q == sum(x[(i,m)] for i in range(n)))
        # car m be used iff there are at least 1 item be packed in car m, so sum(x[(i,m)] for i in range(n)) != 0 
        
        # q = 0 => z[m] = 0
        # q != 0 => z[m] = 1
        model.addConstr(z[m] <= q * M)
        model.addConstr(q <= z[m] * M)

    # objective
    bin_count = sum(z[m] for m in range(k))
    model.setObjective(bin_count, GRB.MINIMIZE)
    
    # Set time limit
    model.setParam('TimeLimit', time_limit)
    
    # Solve the model
    start_time = time.time()
    try: 
        model.optimize()
    except Exception as e:
        print(e)
        result_dict["Result"] = "MO"
        print("MO")
        return n, '-', format(time.time() - start_time, '.6f')
    
    solve_time = time.time() - start_time
    result_dict["Time"] = format(solve_time, '.6f')
    model.write('model_gurobi.lp')
    bins = []
    for j in range(k):
        bins.append([i for i in range(k)  if x[i,j].X > 0.5])
    rot = [int(Ro[i].X) for i in range(n)]
    pos = [(l[i].X, b[i].X) for i in range(n)]
    print(bins)
    print()
   
    print(model.Status)
    if model.Status == GRB.OPTIMAL or model.Status == GRB.TIME_LIMIT:
        print('--------------Solution Found--------------')
        for j in range(k):
            if z[j].X > 0.5:
                print("bin", j+1)
                for i in bins[j]:
                    print(f'put item {i+1} {data["size_item"][i]}with rotation {int(Ro[i].X)} in bin {j+1} at ({l[i].X},{b[i].X})')
                print("-------")
                display_solution((W, H), [data['size_item'][i] for i in bins[j]],[ pos[i] for i in bins[j]], [rot[i] for i in bins[j]])

        # for i in range(n):
        #     print(f'put item {i+1} {data["size_item"][i]}with rotation {int(Ro[i].X)}', end=' ') 
        #     for j in range(k):
        #         if x[i,j].X > 0.5:
        #             print(f'in bin {j+1}', end=' ')
        #     print(f'at ({l[i].X},{b[i].X})')
        
        bin_used = sum(z[m].X for m in range(k))
        result_dict["Number of bins"] = int(bin_used)
        print(f'Number of bin used  :', int(bin_used))
        print('----------------Statistics----------------')
        if model.Status == GRB.OPTIMAL:
            result_dict["Result"] = "SAT"
        else:
            result_dict["Result"] = "TIMEOUT"
        print(f'Time limit          : {time_limit}')
        print(f'Running time        : {solve_time}')
        for j in range(k):
            if z[j].X > 0.5:
                print(j, z[j].X)
        return n, int(bin_used), format(solve_time, '.6f')
    else:
        result_dict["Result"] = "UNSAT"
        print('NO SOLUTIONS')
        return n, '-', format(solve_time, '.6f')


for i in range(1, 10):
    try:
        # Get input file path
        time_limit = int(sys.argv[2])
    except IndexError:
        time_limit = 300
       
    # Create solver
    
    file_path = f'input_data/test.txt'
    print("Reading file: ", file_path.split("/")[-1])
    start = time.time()
    n, n_bins, solver_time = main_solver(file_path, time_limit)
    stop = time.time()

    write_to_xlsx(result_dict)




