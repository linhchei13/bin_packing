from datetime import datetime
import os
import sys
import time
from typing import List, NamedTuple, Tuple
from dataclasses import dataclass, field
from zipfile import BadZipFile

from matplotlib import pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

MAXN = 10007

inf = sys.maxsize

# DECLARATION
N_items = 0
N_bins = 0
bin_used = 0
check_algorithm = False

# BUILD STRUCTURES FOR ITEMS
class Items:
    def __init__(self):
        self.width = 0
        self.height = 0
        self.area = 0
        self.corner_x = 0
        self.corner_y = 0
        self.id = 0
        self.pos_bin = 0
        self.rotated = False

def compare_item_by_longer_side(a: Items):
    if a.height < a.width:
        return a.width
    return a.height 
def compare_item_by_shoter_side(a: Items):
    if a.height > a.width:
        return a.width
    return a.height 
def compare_reset_item(a: Items, b: Items) -> bool:
    return a.id < b.id

def rotate_item(pack: Items):
    pack.rotated = True
    pack.width, pack.height = pack.height, pack.width

# BUILD STRUCTURES FOR BINS
class Free_Rectangles:
    def __init__(self):
        self.corner_x = 0
        self.corner_y = 0
        self.width = 0
        self.height = 0

    def __eq__(self, other):
        return (self.corner_x == other.corner_x and
                self.corner_y == other.corner_y and
                self.width == other.width and
                self.height == other.height)

class Bins:
    def __init__(self):
        self.width = 0
        self.height = 0
        self.area = 0
        self.free_area = 0
        self.id = 0
        self.list_of_free_rec: List[Free_Rectangles] = []
        self.list_of_items: List[Items] = []


# GENERAL PACKING ITEMS FUNCTIONS
# Check if the item fit a specific free_rec
def check_fit_rec(rec: Free_Rectangles, pack: Items, rotated: bool) -> bool:
    # Not rotated case
    if (not rotated) and (pack.width <= rec.width) and (pack.height <= rec.height):
        return True
    # Rotated case
    if rotated and (pack.width <= rec.height) and (pack.height <= rec.width):
        return True
    return False

# Add item to bin with rotated, corner_x, corner_y
def add_item(bin: Bins, pack: Items, rotated: bool, x: int, y: int):
    if rotated:
        rotate_item(pack)
    pack.corner_x = x
    pack.corner_y = y
    bin.list_of_items.append(pack)
    bin.free_area -= pack.area

def compare_ranking_rec_BSS(a: Tuple[int, int], b: Tuple[int, int]) -> bool:
    if a[0] == b[0]:
        return a[1] < b[1]
    return a[0] < b[0]

def score_rec(rec: Free_Rectangles, pack: Items, rotated):
    # Best Short Side: shorter remainder side after insertion is minimized, ties broken with best long
    if rotated:
        score_first = min(rec.width - pack.height, rec.height - pack.width)
        score_second = max(rec.width - pack.height, rec.height - pack.width)
    else:
        score_first = min(rec.width - pack.width, rec.height - pack.height)
        score_second = max(rec.width - pack.width, rec.height - pack.height)
    return (score_first, score_second)

def best_ranking(car, pack):
    rotated = False
    best_rec = Free_Rectangles()
    best_pos = 0
    check_exist = 0
    best_score = (inf, inf)
    
    # Loop to find the best score
    for i, rec in enumerate(car.list_of_free_rec):
        # Not rotate case
        if check_fit_rec(rec, pack, False) and compare_ranking_rec_BSS(score_rec(rec, pack, False), best_score):
            best_score = score_rec(rec, pack, False)
            best_rec = rec
            best_pos = i
            rotated = False
            check_exist = 1
        
        # Rotate case
        if check_fit_rec(rec, pack, True) and compare_ranking_rec_BSS(score_rec(rec, pack, True), best_score):
            best_score = score_rec(rec, pack, True)
            best_rec = rec
            best_pos = i
            rotated = True
            check_exist = 1    
    # Prepare for returning value
    return ((best_rec, best_pos), (rotated, bool(check_exist)))
# PREPARE SOLUTION
def calculate_solution() -> Tuple[int, int]:
    global bin_used
    bin_used = 0
    
    for j in range(N_bins):
        if len(bins[j].list_of_items) > 0:
            bin_used += 1
    
    return (bin_used)

# CHECKING ITEMS AND BINS STATUS
def checking_status(algorithm: bool, item: List[Items]):
    global items_guillotine

    if algorithm == False:
        for i in range(N_items):
            if items_guillotine[i].rotated:
                print(f"Rotate pack {items_guillotine[i].id} and put", end=" ")
            else:
                print(f"Put pack {items_guillotine[i].id}", end=" ")
            print(f"in bin {items_guillotine[i].pos_bin} that (x, y) is "
                  f"({items_guillotine[i].corner_x}, "
                  f"{items_guillotine[i].corner_y})")
    else:
        for i in range(N_items):
            if item[i].rotated:
                print(f"Rotate pack {item[i].id} and put", end=" ")
            else:
                print(f"Put pack {item[i].id}", end=" ")
            print(f"in bin {item[i].pos_bin} that the coordinate (x, y) is "
                  f"({item[i].corner_x}, {item[i].corner_y})")
            

# MAXIMAL RECTANGLES ALGORITHM

# PACKING ITEMS
def spliting_process_maxrec(rec: Free_Rectangles, pack: Items) -> List[Free_Rectangles]:
    list_of_free_rec = []
    new_free_rec = Free_Rectangles()

    right_x     = rec.corner_x + pack.width
    right_y     = rec.corner_y
    right_width = rec.width - pack.width  # Remove the comma at the end
    top_x       = rec.corner_x
    top_y       = rec.corner_y + pack.height
    top_height  = rec.height - pack.height
    split = 1
    # horizontal split if split == 1, otherwise vertical split
    right_height = pack.height if split == 1 else rec.height
    top_width = rec.width if split == 1 else pack.width
    
    if (right_width > 0) and (right_height > 0):
        new_free_rec.corner_x = right_x
        new_free_rec.corner_y = right_y
        new_free_rec.width    = right_width
        new_free_rec.height   = right_height
        list_of_free_rec.append(new_free_rec)
    
    if (top_width > 0) and (top_height > 0):
        new_free_rec = Free_Rectangles()  # Create a new instance
        new_free_rec.corner_x = top_x
        new_free_rec.corner_y = top_y
        new_free_rec.width    = top_width
        new_free_rec.height   = top_height
        list_of_free_rec.append(new_free_rec)
    return list_of_free_rec

def check_intersec_maxrec(rec: Free_Rectangles, pack: Items) -> bool:
    if pack.corner_x >= rec.corner_x + rec.width:   return False
    if pack.corner_y >= rec.corner_y + rec.height:  return False
    if pack.corner_x + pack.width <= rec.corner_x:  return False
    if pack.corner_y + pack.height <= rec.corner_y: return False
    return True

def find_overlap_maxrec(rec: Free_Rectangles, pack: Items) -> Free_Rectangles:
    overlap_rec = Free_Rectangles()
    overlap_rec.corner_x = max(rec.corner_x, pack.corner_x)
    overlap_rec.corner_y = max(rec.corner_y, pack.corner_y)
    overlap_rec.width = min(rec.corner_x + rec.width, pack.corner_x + pack.width) - overlap_rec.corner_x
    overlap_rec.height = min(rec.corner_y + rec.height, pack.corner_y + pack.height) - overlap_rec.corner_y
    return overlap_rec

def split_intersect_maxrec(initial_rec: Free_Rectangles, overlap_rec: Free_Rectangles) -> List[Free_Rectangles]:
    list_of_free_rec = []
    
    # Vertical split to maximize the left corner free_rec
    if overlap_rec.corner_x > initial_rec.corner_x:
        new_free_rec = Free_Rectangles()
        new_free_rec.corner_x = initial_rec.corner_x
        new_free_rec.corner_y = initial_rec.corner_y
        new_free_rec.width = overlap_rec.corner_x - new_free_rec.corner_x
        new_free_rec.height = initial_rec.height
        list_of_free_rec.append(new_free_rec)
    
    # Vertical split to maximize the right corner free_rec
    if overlap_rec.corner_x + overlap_rec.width < initial_rec.corner_x + initial_rec.width:
        new_free_rec = Free_Rectangles()
        new_free_rec.corner_x = overlap_rec.corner_x + overlap_rec.width
        new_free_rec.corner_y = initial_rec.corner_y
        new_free_rec.width = initial_rec.corner_x + initial_rec.width - new_free_rec.corner_x
        new_free_rec.height = initial_rec.height
        list_of_free_rec.append(new_free_rec)
    
    # Horizontal split to maximize the bottom corner free_rec
    if overlap_rec.corner_y > initial_rec.corner_y:
        new_free_rec = Free_Rectangles()
        new_free_rec.corner_x = initial_rec.corner_x
        new_free_rec.corner_y = initial_rec.corner_y
        new_free_rec.width = initial_rec.width
        new_free_rec.height = overlap_rec.corner_y - new_free_rec.corner_y
        list_of_free_rec.append(new_free_rec)
    
    # Horizontal split to maximize the top corner free_rec
    if overlap_rec.corner_y + overlap_rec.height < initial_rec.corner_y + initial_rec.height:
        new_free_rec = Free_Rectangles()
        new_free_rec.corner_x = initial_rec.corner_x
        new_free_rec.corner_y = overlap_rec.corner_y + overlap_rec.height
        new_free_rec.width = initial_rec.width
        new_free_rec.height = initial_rec.corner_y + initial_rec.height - new_free_rec.corner_y
        list_of_free_rec.append(new_free_rec)
    
    return list_of_free_rec

def check_covered_maxrec(rec_covering: Free_Rectangles, rec_covered: Free_Rectangles) -> bool:
    # Not intersect
    if rec_covered.corner_x > rec_covering.corner_x + rec_covering.width:   return False
    if rec_covered.corner_y > rec_covering.corner_y + rec_covering.height:  return False
    if rec_covered.corner_x + rec_covered.width < rec_covering.corner_x:    return False
    if rec_covered.corner_y + rec_covered.height < rec_covering.corner_y:   return False
    
    # Intersect but not fully covered
    if rec_covered.corner_x < rec_covering.corner_x: return False
    if rec_covered.corner_y < rec_covering.corner_y: return False
    if rec_covered.corner_x + rec_covered.width > rec_covering.corner_x + rec_covering.width:   return False
    if rec_covered.corner_y + rec_covered.height > rec_covering.corner_y + rec_covering.height: return False
    
    return True

def remove_covered_rec_maxrec(car: Bins):
    i = 0
    while i < len(car.list_of_free_rec):
        first = car.list_of_free_rec[i]
        j = i + 1
        while j < len(car.list_of_free_rec):
            second = car.list_of_free_rec[j]
            # If rec i cover rec j then delete rec j
            if check_covered_maxrec(first, second):
                car.list_of_free_rec.pop(j)
                continue
            # If rec j cover rec i then delete rec i
            if check_covered_maxrec(second, first):
                car.list_of_free_rec.pop(i)
                i -= 1
                break
            j += 1
        i += 1

def remove_overlap_maxrec(car: Bins, pack: Items):
    i = 0
    while i < len(car.list_of_free_rec):
        rec = car.list_of_free_rec[i]
        if check_intersec_maxrec(rec, pack):
            overlap_rec = find_overlap_maxrec(rec, pack)
            new_rec = split_intersect_maxrec(rec, overlap_rec)
            car.list_of_free_rec.pop(i)
            car.list_of_free_rec.extend(new_rec)
            i -= 1
        i += 1
    remove_covered_rec_maxrec(car)

def insert_item_maxrec(car: Bins, pack: Items) -> bool:
    best_ranking_return = best_ranking(car, pack)
    # If the free_rec which fits the item does not exist
    if not best_ranking_return[1][1]:
        return False
    
    # If the free_rec exists
    pack.pos_bin = car.id
    best_rec = best_ranking_return[0][0]
    best_pos = best_ranking_return[0][1]
    rotated = best_ranking_return[1][0]
    # Add the item into the chosen free_rec
    add_item(car, pack, rotated, best_rec.corner_x, best_rec.corner_y)
    # Replace the used free_rec by the new split rec(s)
    car.list_of_free_rec.pop(best_pos)
    new_rec = spliting_process_maxrec(best_rec, pack)
    car.list_of_free_rec.extend(new_rec)
    # Remove overlap part
    remove_overlap_maxrec(car, pack)
    
    return True

# BIN CHOOSING
def Solve_maxrec(N_items: int, N_bins: int, bin: List[Bins], item: List[Items]):
    for i in range(N_items):
        # Bin First Fit: choose bin that first fit
        for j in range(N_bins):
            if insert_item_maxrec(bin[j], item[i]):
                break
# GUILLLOTINE ALGORITHM


def splitting_process_guillotine(horizontal: bool, rec: Free_Rectangles, pack: Items) -> List[Free_Rectangles]:
    list_of_free_rec = []
    new_free_rec = Free_Rectangles()
    right_x = rec.corner_x + pack.width
    right_y = rec.corner_y
    right_width = rec.width - pack.width
    top_x = rec.corner_x
    top_y = rec.corner_y + pack.height
    top_height = rec.height - pack.height

    right_height = pack.height if horizontal else rec.height
    top_width = rec.width if horizontal else pack.width

    if right_width > 0 and right_height > 0:
        new_free_rec.corner_x = right_x
        new_free_rec.corner_y = right_y
        new_free_rec.width = right_width
        new_free_rec.height = right_height
        list_of_free_rec.append(new_free_rec)
    
    if top_width > 0 and top_height > 0:
        new_free_rec = Free_Rectangles()
        new_free_rec.corner_x = top_x
        new_free_rec.corner_y = top_y
        new_free_rec.width = top_width
        new_free_rec.height = top_height
        list_of_free_rec.append(new_free_rec)
    return list_of_free_rec

def splitting_guillotine(rec: Free_Rectangles, pack: Items) -> List[Free_Rectangles]:
    return splitting_process_guillotine(rec.width <= rec.height, rec, pack)

def merge_rec_guillotine(car: Bins):
    i = 0
    while i < len(car.list_of_free_rec):
        first = car.list_of_free_rec[i]
        check_exist_width = False
        check_exist_height = False
        pos_check_width = 0
        pos_check_height = 0

        for j, second in enumerate(car.list_of_free_rec):
            if j == i:
                continue
            if (first.width == second.width and first.corner_x == second.corner_x and
                second.corner_y == first.corner_y + first.height):
                check_exist_width = True
                pos_check_width = j
                break
            if (first.height == second.height and first.corner_y == second.corner_y and
                second.corner_x == first.corner_x + first.width):
                check_exist_height = True
                pos_check_height = j
                break

        if check_exist_width:
            merged_rec = Free_Rectangles()
            merged_rec.corner_x = first.corner_x
            merged_rec.corner_y = first.corner_y
            merged_rec.width = first.width
            merged_rec.height = first.height + car.list_of_free_rec[pos_check_width].height
            car.list_of_free_rec.pop(pos_check_width)
            if pos_check_width < i:
                i -= 1
            car.list_of_free_rec.pop(i)
            car.list_of_free_rec.append(merged_rec)
            i -= 1
        elif check_exist_height:
            merged_rec = Free_Rectangles()
            merged_rec.corner_x = first.corner_x
            merged_rec.corner_y = first.corner_y
            merged_rec.width = first.width
            merged_rec.height = first.height + car.list_of_free_rec[pos_check_height].height
            car.list_of_free_rec.pop(pos_check_height)
            if pos_check_height < i:
                i -= 1
            car.list_of_free_rec.pop(i)
            car.list_of_free_rec.append(merged_rec)
            i -= 1
        i += 1

def insert_item_guillotine(car: Bins, pack: Items) -> bool:
    best_ranking_return = best_ranking(car, pack)
    
    if not best_ranking_return[1][1]:
        return False
    
    pack.pos_bin = car.id
    best_rec, best_pos = best_ranking_return[0]
    rotated = best_ranking_return[1][0]
    
    add_item(car, pack, rotated, best_rec.corner_x, best_rec.corner_y)
    car.list_of_free_rec.pop(best_pos)
    new_rec = splitting_guillotine(best_rec, pack)
    car.list_of_free_rec.extend(new_rec)
    merge_rec_guillotine(car)
    
    return True

def solve_guillotine(items: List[Items], bins: List[Bins]):
    for item in items:
        for bin in bins:
            if insert_item_guillotine(bin, item):
                break


def enter():
    global N_items, N_bins, items, bins
    N_items = int(input().split()[0])
    N_bins = N_items
    W_bin, H_bin = map(int, input().split())
    bins = [Bins() for _ in range(N_bins)]
    for j in range(N_bins):
        bins[j].width, bins[j].height = W_bin, H_bin
        bins[j].area = bins[j].width * bins[j].height
        bins[j].id = j
        bins[j].free_area = bins[j].area
        first_rec = Free_Rectangles()
        first_rec.width = bins[j].width
        first_rec.height = bins[j].height
        first_rec.corner_x = first_rec.corner_y = 0
        first_rec.area = first_rec.width * first_rec.height
        bins[j].list_of_free_rec.append(first_rec)
    
    items = [Items() for _ in range(N_items)]
    for i in range(N_items):
        input_line = input().split()
        items[i].width, items[i].height = int(input_line[0]), int(input_line[1])
        if items[i].width > items[i].height:
            rotate_item(items[i])
        items[i].area = items[i].width * items[i].height
        items[i].id = i
    items = sorted(items, key=lambda x: (compare_item_by_longer_side(x), compare_item_by_shoter_side(x)), reverse=True)
def reset():
    global items_guillotine, items, bins
    items_guillotine = items.copy()

    for i in range(N_items):
        items[i].corner_x = 0
        items[i].corner_y = 0
    
    for j in range(N_bins):
        bins[j].free_area = bins[j].area
        bins[j].list_of_items.clear()
        bins[j].list_of_free_rec.clear()
        first_rec = Free_Rectangles()
        first_rec.width = bins[j].width
        first_rec.height = bins[j].height
        first_rec.corner_x = first_rec.corner_y = 0
        first_rec.area = first_rec.width * first_rec.height
        bins[j].list_of_free_rec.append(first_rec)

def solve():
    global bin_used, check_algorithm
    enter()
    solve_guillotine(items, bins)
    guillotine_result = calculate_solution()
    reset()
    Solve_maxrec(N_items, N_bins, bins, items)
    maxrec_result = calculate_solution()
    print("Guillotine:", guillotine_result)
    print("Max rect:", maxrec_result)
    if guillotine_result < maxrec_result:
        bin_used = guillotine_result
        check_algorithm = 0
    else:
        bin_used = maxrec_result
        check_algorithm = 1

def print_output():
    print(f"Number of item given: {N_items}")
    print(f"Number of bin given: {N_bins}")
    
    checking_status(check_algorithm, items)
    
    print(f"Number of bin used: {bin_used}")
    # display_solution((bins[0].width, bins[0].height), [(item.width, item.height) for item in bin_used], [(item.corner_x, item.corner_y) for item in items], [item.rotated for item in items])


def display_solution(strip, rectangles, pos_circuits, rotation):
    # define Matplotlib figure and axis
    fig, ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)
    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i],
                                 rectangles[i][0] if not rotation[i] else rectangles[i][1],
                                 rectangles[i][1] if not rotation[i] else rectangles[i][0],
                                 edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    # display plot
    plt.show()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path = 'out/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)
    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")


def main():
    file = ''
    if len(sys.argv) < 2:
        print("Error: No file name provided.")
        file = "input_data/BENG/BENG/BENG01.ins2D"
    else:
        file = sys.argv[1]
    with open(file, 'r') as f:
        sys.stdin = f
        
        start_timing = time.time()
        
        solve()
        
        end_timing = time.time()
        result_dict = {}
        result_dict ={
            "Type": "Heuristic",
            "Data": os.path.basename(sys.argv[1]),
            "Number of items": N_items,
            "Minimize Bin": bin_used,  
            "Solver time": end_timing - start_timing, 
            "Number of variables": "-", 
            "Number of clauses": "-"
        }
        write_to_xlsx(result_dict)
        print_output()
        print("Status: None")
        print("Time limit: None")
        print(f"Running time: {end_timing - start_timing:.20f}")

if __name__ == "__main__":
    main()
