from cplex import Cplex
import sys
import pandas as pd
import os
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime 
from matplotlib import pyplot as plt
W, H = 0, 0
global result_dict
result_dict = {}

def input_data(file_path):
    global W, H
    data = {}
    f = open(file_path,'r')
    
    data['size_item'] = []
    data['size_bin'] = []
    n = int(f.readline())
    W, H = map(int, f.readline().split())
    data['size_bin'].append([W, H])
    for i in range(1,n+1):
        line = f.readline().split()
        data['size_item'].append([int(line[0]),int(line[1])])
    
    return n,data,W,H


def display_solution(strip, rectangles, pos_circuits, rotation):
    # define Matplotlib figure and axis
    fig, ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i],
                                 rectangles[i][0] if not rotation[i] else rectangles[i][1],
                                 rectangles[i][1] if not rotation[i] else rectangles[i][0],
                                 edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    # display plot
    plt.show()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path = 'out/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)
    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")

def main_solver(file_path, time_limit):
    n,data,W,H = input_data(file_path)
    k = n
    print('Number of items:',n)
    print('Size of bin:',W,H)
    
    # Create CPLEX model
    model = Cplex()
    model.objective.set_sense(model.objective.sense.minimize)
    
    # Set CPLEX parameters
    model.parameters.timelimit.set(time_limit)
    model.parameters.emphasis.mip.set(1)  # Focus on feasibility
    model.parameters.mip.tolerances.mipgap.set(0.05)  # 5% MIP gap
    model.parameters.mip.tolerances.integrality.set(1e-6)  # Tighter integrality tolerance
    model.parameters.simplex.tolerances.feasibility.set(1e-6)  # Tighter feasibility tolerance
    
    # Create variables
    M = max(W, H) * 3  # Big-M value

    # Variable names
    x_names = []
    ro_names = []
    l_names = []
    r_names = []
    t_names = []
    b_names = []
    
    # Create variables
    for i in range(n):
        ro_names.append(f'Ro[{i}]')
        l_names.append(f'l[{i}]')
        r_names.append(f'r[{i}]')
        t_names.append(f't[{i}]')
        b_names.append(f'b[{i}]')
        
        for m in range(k):
            x_names.append(f'x[{i},{m}]')
    
    # Add variables to model
    model.variables.add(
        types=[model.variables.type.binary] * len(ro_names),
        names=ro_names
    )
    
    model.variables.add(
        lb=[0] * len(l_names),
        ub=[W] * len(l_names),
        types=[model.variables.type.continuous] * len(l_names),
        names=l_names
    )
    
    model.variables.add(
        lb=[0] * len(r_names),
        ub=[W] * len(r_names),
        types=[model.variables.type.continuous] * len(r_names),
        names=r_names
    )
    
    model.variables.add(
        lb=[0] * len(t_names),
        ub=[H] * len(t_names),
        types=[model.variables.type.continuous] * len(t_names),
        names=t_names
    )
    
    model.variables.add(
        lb=[0] * len(b_names),
        ub=[H] * len(b_names),
        types=[model.variables.type.continuous] * len(b_names),
        names=b_names
    )
    
    model.variables.add(
        types=[model.variables.type.binary] * len(x_names),
        names=x_names
    )
    
    # Add bin usage variables
    z_names = [f'z[{m}]' for m in range(k)]
    model.variables.add(
        types=[model.variables.type.binary] * len(z_names),
        names=z_names
    )
    
    # Add constraints for item dimensions
    # Rotation constraints for CPLEX
    for i in range(n):
    # r[i] = (1-Ro[i]) * data['size_item'][i][0] + Ro[i] * data['size_item'][i][1] + l[i]
        model.linear_constraints.add(
            lin_expr=[[
                [f"r[{i}]", f"Ro[{i}]", f"l[{i}]"], 
                [1, -data['size_item'][i][1] + data['size_item'][i][0], -1]
            ]],
            senses=["E"],
            rhs=[data['size_item'][i][0]]
        )
        
        # t[i] = (1-Ro[i]) * data['size_item'][i][1] + Ro[i] * data['size_item'][i][0] + b[i]
        model.linear_constraints.add(
            lin_expr=[[
                [f"t[{i}]", f"Ro[{i}]", f"b[{i}]"], 
                [1, -data['size_item'][i][0] + data['size_item'][i][1], -1]
            ]],
            senses=["E"],
            rhs=[data['size_item'][i][1]]
        )
        
        # IMPROVED BOUNDARY CONSTRAINTS - Ensure items stay within bin boundaries
        for m in range(k):
            # When item i is in bin m:
            # 1. left edge in bounds: l[i] >= 0
            # 2. bottom edge in bounds: b[i] >= 0
            # 3. right edge in bounds: r[i] <= W
            # 4. top edge in bounds: t[i] <= H
            
            # l[i] >= 0 when x[i,m] = 1
            model.indicator_constraints.add(
                indvar=f"x[{i},{m}]",
                complemented=0,
                lin_expr=[[f"l[{i}]"], [1]],
                sense="G",
                rhs=0
            )
            
            # b[i] >= 0 when x[i,m] = 1
            model.indicator_constraints.add(
                indvar=f"x[{i},{m}]",
                complemented=0,
                lin_expr=[[f"b[{i}]"], [1]],
                sense="G",
                rhs=0
            )
            
            # r[i] <= W when x[i,m] = 1
            model.indicator_constraints.add(
                indvar=f"x[{i},{m}]",
                complemented=0,
                lin_expr=[[f"r[{i}]"], [1]],
                sense="L",
                rhs=W
            )
            
            # t[i] <= H when x[i,m] = 1
            model.indicator_constraints.add(
                indvar=f"x[{i},{m}]",
                complemented=0,
                lin_expr=[[f"t[{i}]"], [1]],
                sense="L",
                rhs=H
            )
    
    # Each item must be in exactly one bin
    for i in range(n):
        model.linear_constraints.add(
            lin_expr=[[
                [f"x[{i},{m}]" for m in range(k)],
                [1] * k
            ]],
            senses=["E"],
            rhs=[1]
        )
    
    # Non-overlap constraints
    for i in range(n - 1):
        for j in range(i + 1, n):
            for m in range(k):
                # Create e variable for "both items are in bin m"
                e_name = f'e[{i},{j},{m}]'
                model.variables.add(
                    types=[model.variables.type.binary],
                    names=[e_name]
                )
                
                # e >= x[i,m] + x[j,m] - 1
                model.linear_constraints.add(
                    lin_expr=[[
                        [e_name, f"x[{i},{m}]", f"x[{j},{m}]"],
                        [1, -1, -1]
                    ]],
                    senses=["G"],
                    rhs=[-1]
                )
                
                # e <= x[i,m]
                model.linear_constraints.add(
                    lin_expr=[[
                        [e_name, f"x[{i},{m}]"],
                        [1, -1]
                    ]],
                    senses=["L"],
                    rhs=[0]
                )
                
                # e <= x[j,m]
                model.linear_constraints.add(
                    lin_expr=[[
                        [e_name, f"x[{j},{m}]"],
                        [1, -1]
                    ]],
                    senses=["L"],
                    rhs=[0]
                )
                
                # Create binary variables for the four non-overlap conditions
                c_names = [f'c1[{i},{j},{m}]', f'c2[{i},{j},{m}]', f'c3[{i},{j},{m}]', f'c4[{i},{j},{m}]']
                model.variables.add(
                    types=[model.variables.type.binary] * 4,
                    names=c_names
                )
                
                # Constraints representing the four non-overlap conditions
                # c1: r[i] <= l[j]
                model.linear_constraints.add(
                    lin_expr=[[
                        [f"r[{i}]", f"l[{j}]", c_names[0]],
                        [1, -1, M]
                    ]],
                    senses=["L"],
                    rhs=[M]
                )
                
                # c2: r[j] <= l[i]
                model.linear_constraints.add(
                    lin_expr=[[
                        [f"r[{j}]", f"l[{i}]", c_names[1]],
                        [1, -1, M]
                    ]],
                    senses=["L"],
                    rhs=[M]
                )
                
                # c3: t[i] <= b[j]
                model.linear_constraints.add(
                    lin_expr=[[
                        [f"t[{i}]", f"b[{j}]", c_names[2]],
                        [1, -1, M]
                    ]],
                    senses=["L"],
                    rhs=[M]
                )
                
                # c4: t[j] <= b[i]
                model.linear_constraints.add(
                    lin_expr=[[
                        [f"t[{j}]", f"b[{i}]", c_names[3]],
                        [1, -1, M]
                    ]],
                    senses=["L"],
                    rhs=[M]
                )
                
                # If both items are in bin m, at least one non-overlap constraint must be active
                model.linear_constraints.add(
                    lin_expr=[[
                        c_names + [e_name],
                        [1, 1, 1, 1, -M]
                    ]],
                    senses=["G"],
                    rhs=[1 - M]
                )
                
                # All constraints must be zero if e is zero
                model.linear_constraints.add(
                    lin_expr=[[
                        c_names + [e_name],
                        [1, 1, 1, 1, -M]
                    ]],
                    senses=["L"],
                    rhs=[0]
                )
    
    # Bin usage constraints
    for m in range(k):
        q_name = f'q[{m}]'
        model.variables.add(
            lb=[0],
            ub=[n],
            types=[model.variables.type.integer],
            names=[q_name]
        )
        
        # q = sum of x[i,m] for all i
        model.linear_constraints.add(
            lin_expr=[[
                [q_name] + [f"x[{i},{m}]" for i in range(n)],
                [1] + [-1] * n
            ]],
            senses=["E"],
            rhs=[0]
        )
        
        # z[m] = 1 if q > 0 (bin is used)
        # z[m] <= q * M
        model.linear_constraints.add(
            lin_expr=[[
                [f"z[{m}]", q_name],
                [1, -1]
            ]],
            senses=["L"],
            rhs=[0]
        )
        
        # q <= z[m] * M
        model.linear_constraints.add(
            lin_expr=[[
                [q_name, f"z[{m}]"],
                [1, -M]
            ]],
            senses=["L"],
            rhs=[0]
        )
    
    # Add symmetry breaking constraint: use bins in order
    for m in range(1, k):
        model.linear_constraints.add(
            lin_expr=[[
                [f"z[{m}]", f"z[{m-1}]"],
                [1, -1]
            ]],
            senses=["L"],
            rhs=[0]
        )
    
    # Set objective to minimize number of bins used
    # Only include z variables in the objective function with coefficient 1
    model.objective.set_linear([(z_name, 1) for z_name in z_names])
    
    # Make sure no other variables accidentally get into objective
    # Clear any potential zero coefficients from other variables
    for name in ro_names + l_names + r_names + t_names + b_names:
        if name in model.objective.get_linear():
            model.objective.set_linear(name, 0)
            
    model.write('model_cplex.lp')
    # Solve the model
    print('--------------Solving--------------')
    model.solve()
    print('--------------Solved--------------')
    status = model.solution.get_status()
    
    # Get solution status
    solve_time = model.get_time()/1000
    for j in range(k):
        print(f"Z{j}",model.solution.get_values(f'z[{j}]'))
        for i in range(n):
            print(f"X{i},{j}",model.solution.get_values(f'x[{i},{j}]'))

    bins = []
    for j in range(k):
        bins.append([i for i in range(n) if model.solution.get_values(f'x[{i},{j}]') > 0.5])
    
    rot = []
    pos = []
    
    global result_dict
    result_dict = {
        "Type": "CPLEX",
        'Problem': file_path.split("/")[-1],
        "Number of items": n,
        "Width": W,
        "Height": H,
        "Number of bins": 0,
        "Time": format(solve_time, '.6f'),
        "Result": "SAT",
    }
    
    if status in [model.solution.status.optimal, model.solution.status.MIP_optimal,
                  model.solution.status.feasible, model.solution.status.MIP_feasible]:
        print('--------------Solution Found--------------')
        for i in range(n):
            print('Item', i+1, 'Rotation:', int(model.solution.get_values(f'Ro[{i}]')),
                  "size:", data['size_item'][i],
                  'Position:', [model.solution.get_values(f"l[{i}]"), model.solution.get_values(f"b[{i}]")],
                  "Top-Right:", [model.solution.get_values(f"r[{i}]"), model.solution.get_values(f"t[{i}]")])
            ro_val = int(model.solution.get_values(f'Ro[{i}]'))
            rot.append(ro_val)
            pos.append((model.solution.get_values(f"l[{i}]"), model.solution.get_values(f"b[{i}]")))
        
        for j in range(k):
            if model.solution.get_values(f'z[{j}]') > 0.5:
                print(f"Bin {j+1}:")
                for i in range(n):
                    if model.solution.get_values(f'x[{i},{j}]') > 0.5:
                        print(f'  Put item {i+1} {data["size_item"][i]} with rotation {rot[i]} at ({pos[i][0]},{pos[i][1]})')
                print("-------")
                # Uncomment to display solution
                display_solution((W, H), [data['size_item'][i] for i in bins[j]],
                               [pos[i] for i in bins[j]], [rot[i] for i in bins[j]])
        
        bin_used = sum(model.solution.get_values(f'z[{m}]') for m in range(k))
        result_dict["Number of bins"] = int(bin_used)
        print(f'Number of bins used: {int(bin_used)}')
        print('----------------Statistics----------------')
        
        if status in [model.solution.status.optimal, model.solution.status.MIP_optimal]:
            result_dict["Result"] = "OPTIMAL"
            print('Status: OPTIMAL')
        else:
            result_dict["Result"] = "FEASIBLE"
            print('Status: FEASIBLE')
            
        print(f'Time limit: {time_limit}')
        print(f'Running time: {solve_time}')               
        return n, int(bin_used), format(solve_time, '.6f')
    else:
        result_dict["Result"] = "UNSAT"
        print('NO SOLUTIONS')
        return n, '-', format(solve_time, '.6f')


if __name__ == '__main__':
    
        try:
            # Get time limit
            time_limit = int(sys.argv[1]) if len(sys.argv) > 1 else 600
        except IndexError:
            time_limit = 600
           
        # Create solver
        file_path = f'input_data/class/cl_020_02.txt'
        print("Reading file: ", file_path.split("/")[-1])
        start = time.time()
        n, n_bins, solver_time = main_solver(file_path, time_limit)
        stop = time.time()
        
        result_dict["Real time"] = format(stop - start, '.6f')
        write_to_xlsx(result_dict)
