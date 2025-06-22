from itertools import chain, combinations
import math
import os
from threading import Timer

import pandas as pd
from pysat.formula import CNF
import matplotlib.pyplot as plt

from pysat.solvers import Glucose3
from threading import Timer
import datetime
import pandas as pd
import os
import sys
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

class TimeoutException(Exception): pass
# Initialize the CNF formula
n_items = 0
W, H = 0, 0
items = []
time_out = 3000
time_limit_solver = 500
#read file
def read_input():
    global W, H, items, n_items
    n_items = int(input().split()[0])
    W, H = map(int, input().split())
    for i in range(n_items):
        list = input().split() 
        items.append([int(list[0]), int(list[1])])

def positive_range(end):
    if (end < 0):
        return []
    return range(end)

def display_solution(strip, rectangles, pos_circuits, rotation):
    # define Matplotlib figure and axis
    fig, ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i],
                                 rectangles[i][0] if not rotation[i] else rectangles[i][1],
                                 rectangles[i][1] if not rotation[i] else rectangles[i][0],
                                 edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    # display plot
    plt.show()

def generate_all_clauses(rectangles, n, W, H):
# Define the variables
    height = H
    width = n * W
    clauses = []
    variables = {}
    id_variables = 1
    start = time.time()
    for i in range(len(rectangles)):
        for j in range(len(rectangles)):
            if i != j:
                variables[f"lr{i + 1},{j + 1}"] = id_variables  # lri,rj
                id_variables += 1
                variables[f"ud{i + 1},{j + 1}"] = id_variables  # uri,rj
                id_variables += 1
        for e in range(width):
            variables[f"px{i + 1},{e}"] = id_variables  # pxi,e
            id_variables += 1
        for f in range(height):
            variables[f"py{i + 1},{f}"] = id_variables  # pyi,f
            id_variables += 1

    # Rotated variables
    for i in range(len(rectangles)):
        variables[f"r{i + 1}"] = id_variables
        id_variables += 1
    
    # Add the 2-literal axiom clauses
    for i in range(len(rectangles)):
        for e in range(width - 1):  # -1 because we're using e+1 in the clause
            clauses.append([-variables[f"px{i + 1},{e}"],
                        variables[f"px{i + 1},{e + 1}"]])
        for f in range(height - 1):  # -1 because we're using f+1 in the clause
            clauses.append([-variables[f"py{i + 1},{f}"],
                        variables[f"py{i + 1},{f + 1}"]])
    # Add non-overlapping constraints

    def non_overlapping(rotated, i, j):
        if not rotated:
            i_width = rectangles[i][0]
            i_height = rectangles[i][1]
            j_width = rectangles[j][0]
            j_height = rectangles[j][1]
            i_rotation = variables[f"r{i + 1}"]
            j_rotation = variables[f"r{j + 1}"]
        else:
            i_width = rectangles[i][1]
            i_height = rectangles[i][0]
            j_width = rectangles[j][1]
            j_height = rectangles[j][0]
            i_rotation = -variables[f"r{i + 1}"]
            j_rotation = -variables[f"r{j + 1}"]

        # lri,j v lrj,i v udi,j v udj,i
        four_literal = []
        four_literal.append(variables[f"lr{i + 1},{j + 1}"])
        four_literal.append(variables[f"lr{j + 1},{i + 1}"])
        four_literal.append(variables[f"ud{i + 1},{j + 1}"])
        four_literal.append(variables[f"ud{j + 1},{i + 1}"])

        clauses.append(four_literal + [i_rotation])
        clauses.append(four_literal + [j_rotation])

        # ¬lri, j ∨ ¬pxj, e
        
        for e in range(min(width, i_width)):
                clauses.append([i_rotation,
                            -variables[f"lr{i + 1},{j + 1}"],
                            -variables[f"px{j + 1},{e}"]])
        # ¬lrj,i ∨ ¬pxi,e
        
        for e in range(min(width, j_width)):
                clauses.append([j_rotation,
                            -variables[f"lr{j + 1},{i + 1}"],
                            -variables[f"px{i + 1},{e}"]])
       
        for f in range(min(height, i_height)):
                clauses.append([i_rotation,
                            -variables[f"ud{i + 1},{j + 1}"],
                            -variables[f"py{j + 1},{f}"]])
        # ¬udj, i ∨ ¬pyi, f,
        
        for f in range(min(height, j_height)):
                clauses.append([j_rotation,
                            -variables[f"ud{j + 1},{i + 1}"],
                            -variables[f"py{i + 1},{f}"]])

        for e in positive_range(width - i_width):
            # ¬lri,j ∨ ¬pxj,e+wi ∨ pxi,e
        
            clauses.append([i_rotation,
                        -variables[f"lr{i + 1},{j + 1}"],
                        variables[f"px{i + 1},{e}"],
                        -variables[f"px{j + 1},{e + i_width}"]])

        for e in positive_range(width - j_width):
            # ¬lrj,i ∨ ¬pxi,e+wj ∨ pxj,e
                    clauses.append([j_rotation,
                                -variables[f"lr{j + 1},{i + 1}"],
                                variables[f"px{j + 1},{e}"],
                                -variables[f"px{i + 1},{e + j_width}"]])

        for f in positive_range(height - i_height):
            # udi,j ∨ ¬pyj,f+hi ∨ pxi,e
                clauses.append([i_rotation,
                            -variables[f"ud{i + 1},{j + 1}"],
                            variables[f"py{i + 1},{f}"],
                            -variables[f"py{j + 1},{f + i_height}"]])
        for f in positive_range(height - j_height):
            # ¬udj,i ∨ ¬pyi,f+hj ∨ pxj,f
            
            clauses.append([j_rotation,
                        -variables[f"ud{j + 1},{i + 1}"],
                        variables[f"py{j + 1},{f}"],
                        -variables[f"py{i + 1},{f + j_height}"]])

    for i in range(len(rectangles)):
        for j in range(i + 1, len(rectangles)):
                non_overlapping(False, i, j)
                non_overlapping(True, i, j)

    # # Domain encoding to ensure every rectangle stays inside strip's boundary
    for i in range(len(rectangles)):
            if rectangles[i][0] > width: #if rectangle[i]'s width larger than strip's width, it has to be rotated
                clauses.append([variables[f"r{i + 1}"]])
            else:
                clauses.append([variables[f"r{i + 1}"],
                                    variables[f"px{i + 1},{width - rectangles[i][0]}"]])
       
            if rectangles[i][1] > height:
                clauses.append([variables[f"r{i + 1}"]])
            else:
                clauses.append([variables[f"r{i + 1}"],
                            variables[f"py{i + 1},{height - rectangles[i][1]}"]])

            # Rotated
            if rectangles[i][1] > width:
                clauses.append([-variables[f"r{i + 1}"]])
            else:
                clauses.append([-variables[f"r{i + 1}"],
                                    variables[f"px{i + 1},{width - rectangles[i][1]}"]])
            if rectangles[i][0] > height:
                clauses.append([-variables[f"r{i + 1}"]])
            else:
                clauses.append([-variables[f"r{i + 1}"],
                                variables[f"py{i + 1},{height - rectangles[i][0]}"]])
    
    for k in range(1, n):
         for i in range(len(rectangles)):
            # Not rotated
            w = rectangles[i][0]
            clauses.append([variables[f"r{i + 1}"], variables[f"px{i + 1},{k * W - w}"],
                        -variables[f"px{i + 1},{k * W - 1}"]])
            clauses.append([variables[f"r{i + 1}"], -variables[f"px{i + 1},{k * W - w}"],
                        variables[f"px{i + 1},{k * W - 1}"]])
            
            # Rotated
            w = rectangles[i][1]
            clauses.append([-variables[f"r{i + 1}"], variables[f"px{i + 1},{k * W - w}"],
                        -variables[f"px{i + 1},{k * W - 1}"]])
            clauses.append([-variables[f"r{i + 1}"], -variables[f"px{i + 1},{k * W - w}"],
                        variables[f"px{i + 1},{k * W - 1}"]])
    stop = time.time()
    print("Encoding Time:", format(stop-start, ".6f"))        
    return clauses, variables

def solve_bpp(rectangles, n_items, W, H):
    clauses, variables = generate_all_clauses(items, n_items, W, H)
    print("Variables:", len(variables))
    print("Clauses:", len(clauses))
    width = W * n_items
    height = H
    solver = Glucose3(use_timer=True)
    sat_status = False
    def interrupt(solver):
        solver.interrupt()
        
    for clause in clauses:
        solver.add_clause(clause)
    
    timer = Timer(180, interrupt, [solver])
    timer.start()
    start = time.time()
    try: 
        sat_status = solver.solve_limited(expect_interrupt=True)
        if sat_status:
            pos = [[0 for i in range(2)] for j in range(len(rectangles))]
            rotation = []
            model = solver.get_model()
            solver_time = format(time.time() - start, ".6f")
            print("Solver time:", solver_time)
            print("SAT")
            result = {}
            for var in model:
                if var > 0:
                    result[list(variables.keys())[list(variables.values()).index(var)]] = True
                else:
                    result[list(variables.keys())[list(variables.values()).index(-var)]] = False
            print("Decode")
            for i in range(len(rectangles)):
                rotation.append(result[f"r{i + 1}"])
                for e in range(width - 1):
                    if result[f"px{i + 1},{e}"] == False and result[f"px{i + 1},{e + 1}"] == True:
                        pos[i][0] = e + 1
                    if e == 0 and result[f"px{i + 1},{e}"] == True:
                        pos[i][0] = 0
                for f in range(height - 1):
                    if result[f"py{i + 1},{f}"] == False and result[f"py{i + 1},{f + 1}"] == True:
                        pos[i][1] = f + 1
                    if f == 0 and result[f"py{i + 1},{f}"] == True:
                        pos[i][1] = 0  
            timer.cancel()
            return "sat", pos, rotation, solver_time, len(variables), len(clauses)

        else:
            timer.cancel()
            print("unsat")
            return("unsat")
    except TimeoutException:
        print("Timeout")
        return("timeout")
    
    
def BPP(W, H, items, n):
    items_area = [i[0] * i[1] for i in items]
    print("Items area", sum(items_area))
    bin_area = W * H
    lower_bound = math.ceil(sum(items_area) / bin_area)

    for k in range(27, n + 1):
        print(f"Trying with {k} bins")
        result = solve_bpp(items, k, W, H)
        
        if result[0] == "sat":
                print(f"Solution found with {k} bins")
                position = result[1]
                bins_used = [[i for i in range(n) if position[i][0] // W == j] for j in range(k)]
                rotation = result[2]
                solver_time = result[3]
                num_variables = result[4]
                num_clauses = result[5]
                for j in range(k):
                    for i in range(n):
                        if position[i][0] // W == j:
                            position[i][0] = position[i][0] - j * W
                return bins_used, position, rotation, solver_time, num_variables, num_clauses
        

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path = 'out/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)
    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")

def export_to_xlsx(bpp_result, filepath, time):
    result_dict = {}
    if bpp_result is None:
        result_dict = {
            "Type": "Stacking no symmetry",
            "Dataset": filepath.split("/")[-1],
            "Number of items": n_items,
            "Result": "UNSAT",
            "Minimize Bin": "N/A",  
            "Solver time": "-", 
            "Real time": time,
            "Number of variables": "-", 
            "Number of clauses": "-"}
    else:
        bins, pos, rota, solver_time, num_variables, num_clauses = bpp_result
        result_dict = {
            "Type": "Stacking no symmetry",
            "Dataset": filepath.split("/")[-1],
            "Number of items": n_items,
            "Minimize Bin": len(bins),  
            "Solver time": solver_time, 
            "Real time": time,
            "Number of variables": num_variables, 
            "Number of clauses": num_clauses}
    write_to_xlsx(result_dict)
    

def print_solution(bpp_result):
    if bpp_result is None:
        print("No solution found")
    else:
        bins, pos, rotation, solver_time, num_variables, num_clauses = bpp_result
        for i in range(len(bins)):
            print("Bin", i + 1, "contains items", [(j + 1) for j in bins[i]])
            for j in bins[i]:
                if rotation[j]:
                    print("Rotated item", j + 1, items[j], "at position", pos[j])
                else:
                    print("Item", j + 1, items[j], "at position", pos[j])
            display_solution((W, H), [items[j] for j in bins[i]], [pos[j] for j in bins[i]], [rotation[j] for j in bins[i]])
        print("--------------------")
        print("Solution found with", len(bins), "bins")
        print(f"Solver time: {solver_time} seconds")
        print("Number of variables:", num_variables)
        print("Number of clauses:", num_clauses)
    

    # read input file

if __name__ == "__main__":
    filepath = f"input_data/Class/cl_080_03.txt"
    with open(filepath, 'r') as f:
        sys.stdin = f
        items = []
        print(f"Reading file {filepath.split('/')[-1]}")
        read_input()
        start = time.time()
        bpp_result = BPP(W, H, items, n_items)
        stop = time.time()
        print("Time:", format(stop-start, ".6f"))
        export_to_xlsx(bpp_result, filepath, format(stop-start, ".3f"))
        # print_solution(bpp_result)
    

