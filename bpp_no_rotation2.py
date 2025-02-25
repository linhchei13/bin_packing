from itertools import chain, combinations
import math
from threading import Timer

from pysat.formula import CNF
from pysat.solvers import Glucose3

import matplotlib.pyplot as plt
from threading import Timer
import datetime
import pandas as pd
import os
import sys
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

class TimeoutException(Exception): pass

#read file
def read_input():
    global W, H, items, n_items
    n_items = int(input().split()[0])
    W, H = map(int, input().split())
    for i in range(n_items):
        items.append(list(map(int, input().split())))
def read_file_instance(filepath):
    f = open(filepath)
    return f.read().splitlines()

def positive_range(end):
    if (end < 0):
        return []
    return range(end)

def display_solution(strip, rectangles, pos_circuits):
    # define Matplotlib figure and axis
    fig, ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    # display plot
    plt.show()

def OPP(rectangles, max_bins, W, H):
# Define the variables
    cnf = CNF()
    variables = {}
    counter = 1
    width = W
    height = H
    # find max height and width rectangles for largest rectangle symmetry breaking
    max_height = max([int(rectangle[1]) for rectangle in rectangles])
    max_width = max([int(rectangle[0]) for rectangle in rectangles])

    for i in range(len(rectangles)):
        for j in range(max_bins):
            variables[f"x{i + 1},{j + 1}"] = counter 
            counter += 1

    # create lr, ud variables
    for i in range(len(rectangles)):
        for j in range(len(rectangles)):
            variables[f"lr{i + 1},{j + 1}"] = counter  # lri,rj
            counter += 1
            variables[f"ud{i + 1},{j + 1}"] = counter  # uri,rj
            counter += 1
        for e in positive_range(width - rectangles[i][0] + 1):
            variables[f"px{i + 1},{e}"] = counter  # pxi,e
            counter += 1
        for f in positive_range(height - rectangles[i][1] + 1):
            variables[f"py{i + 1},{f}"] = counter  # pyi,f
            counter += 1
     # Exactly one bin for each item
    for i in range(len(rectangles)):
        cnf.append([variables[f"x{i + 1},{j + 1}"] for j in range(max_bins)])
        for j1 in range(max_bins):
            for j2 in range(j1 + 1, max_bins):
                cnf.append([-variables[f"x{i + 1},{j1 + 1}"], -variables[f"x{i + 1},{j2 + 1}"]])

    # Add the 2-literal axiom clauses (order constraint)
    for i in range(len(rectangles)):
       # ¬pxi,e ∨ pxi,e+1
        for e in range(width - rectangles[i][0]):  # -1 because we're using e+1 in the clause
            cnf.append([-variables[f"px{i + 1},{e}"],
                        variables[f"px{i + 1},{e + 1}"]])
        #  ¬pyi,f ∨ pxi,f+1
        for f in range(height - rectangles[i][1]):  # -1 because we're using f+1 in the clause
            cnf.append([-variables[f"py{i + 1},{f}"],
                        variables[f"py{i + 1},{f + 1}"]])


    # Add the 3-literal non-overlapping constraints
    def non_overlapping(i, j, h1, h2, v1, v2, b):
        i_width = rectangles[i][0]
        i_height = rectangles[i][1]
        j_width = rectangles[j][0]
        j_height = rectangles[j][1]
        bin_cnf = [-variables[f"x{i + 1},{b + 1}"], -variables[f"x{j + 1},{b + 1}"]]
        # lri, j ∨ lrj, i ∨ udi, j ∨ udj, i
        four_literal = []
        if h1: four_literal.append(variables[f"lr{i + 1},{j + 1}"])
        if h2: four_literal.append(variables[f"lr{j + 1},{i + 1}"])
        if v1: four_literal.append(variables[f"ud{i + 1},{j + 1}"])
        if v2: four_literal.append(variables[f"ud{j + 1},{i + 1}"])
        cnf.append(four_literal + bin_cnf)

        # ¬lri, j ∨ ¬pxj, e
        if h1:
            for e in range(i_width):
                if f"px{j + 1},{e}" in variables:
                    cnf.append(bin_cnf + [-variables[f"lr{i + 1},{j + 1}"],
                                -variables[f"px{j + 1},{e}"]])
        # ¬lrj,i ∨ ¬pxi,e
        if h2:
            for e in range(j_width):
                if f"px{i + 1},{e}" in variables:
                    cnf.append(bin_cnf +[-variables[f"lr{j + 1},{i + 1}"],
                                -variables[f"px{i + 1},{e}"]])
        # ¬udi,j ∨ ¬pyj,f
        if v1:
            for f in range(i_height):
                if f"py{j + 1},{f}" in variables:
                    cnf.append(bin_cnf +[-variables[f"ud{i + 1},{j + 1}"],
                                -variables[f"py{j + 1},{f}"]])
        # ¬udj, i ∨ ¬pyi, f,
        if v2:
            for f in range(j_height):
                if f"py{i + 1},{f}" in variables:
                    cnf.append(bin_cnf +[-variables[f"ud{j + 1},{i + 1}"],
                                -variables[f"py{i + 1},{f}"]])

        for e in positive_range(width - i_width):
            # ¬lri,j ∨ ¬pxj,e+wi ∨ pxi,e
            if h1:
                if f"px{j + 1},{e + i_width}" in variables:
                    cnf.append(bin_cnf +[-variables[f"lr{i + 1},{j + 1}"],
                                variables[f"px{i + 1},{e}"],
                                -variables[f"px{j + 1},{e + i_width}"]])
            # ¬lrj,i ∨ ¬pxi,e+wj ∨ pxj,e
            if h2:
                if f"px{i + 1},{e + j_width}" in variables:
                    cnf.append(bin_cnf +[-variables[f"lr{j + 1},{i + 1}"],
                                variables[f"px{j + 1},{e}"],
                                -variables[f"px{i + 1},{e + j_width}"]])

        for f in positive_range(height - i_height):
            # udi,j ∨ ¬pyj,f+hi ∨ pxi,e
            if v1:
                if f"py{j + 1},{f + i_height}" in variables:
                    cnf.append(bin_cnf +[-variables[f"ud{i + 1},{j + 1}"],
                                variables[f"py{i + 1},{f}"],
                                -variables[f"py{j + 1},{f + i_height}"]])
            # ¬udj,i ∨ ¬pyi,f+hj ∨ pxj,f
            if v2:
                if f"py{i + 1},{f + j_height}" in variables:
                    cnf.append(bin_cnf +[-variables[f"ud{j + 1},{i + 1}"],
                                variables[f"py{j + 1},{f}"],
                                -variables[f"py{i + 1},{f + j_height}"]])
    for b in range(max_bins):
        for i in range(len(rectangles)):
            for j in range(i + 1, len(rectangles)):
                # lri,j ∨ lrj,i ∨ udi,j ∨ udj,i
                #Large-rectangles horizontal
                if rectangles[i][0] + rectangles[j][0] > width:
                    non_overlapping(i, j, False, False, True, True, b)

                #Large-rectangles vertical
                if rectangles[i][1] + rectangles[j][1] > height:
                    non_overlapping(i, j, True, True, False, False, b)

                #Same-sized rectangles
                elif rectangles[i] == rectangles[j]:
                    non_overlapping(i, j, True, False, True, True, b)
                #
                #largest width rectangle
                elif rectangles[i][0] == max_width and rectangles[j][0] > (width - max_width) / 2:
                    non_overlapping(i, j, False, True, True, True, b)
                #
                #largest height rectangle
                elif rectangles[i][1] == max_height and rectangles[j][1] > (height - max_height) / 2:
                    non_overlapping(i, j, True, True, False, True, b)

            #normal rectangles
                else:
                    non_overlapping(i, j, True, True, True, True, b)

    # Domain encoding for px and py: 0 <= x <= width and 0 <= y <= height
    # equal to: px(i, W-wi) ^ !px(i,-1) and py(i, H-hi) ^ !py(i,-1)
    for b in range(max_bins):
        for i in range(len(rectangles)):
            cnf.append([-variables[f"x{i + 1},{b + 1}"], variables[f"px{i + 1},{width - rectangles[i][0]}"]]) # px(i, W-wi)
            cnf.append([-variables[f"x{i + 1},{b + 1}"],variables[f"py{i + 1},{height - rectangles[i][1]}"]])  # py(i, H-hi)
            
    # Solve the SAT problem
    start = time.time()
    def solver_interrupt(solver):
        print("Timeout")
        solver.interrupt()
    solver = Glucose3(use_timer=True)

    solver.append_formula(cnf)
    timer = Timer(200, solver_interrupt, [solver])
    timer.start()
    try:
        sat_status = solver.solve_limited(expect_interrupt=True)
        if sat_status:
            pos = [[0 for i in range(2)] for j in range(len(rectangles))]
            model = solver.get_model()
            print("SAT")
            result = {}
            timer.cancel()
            for var in model:
                if var > 0:
                    result[list(variables.keys())[list(variables.values()).index(var)]] = True
                else:
                    result[list(variables.keys())[list(variables.values()).index(-var)]] = False
            stop = time.time()
            # from SAT result, decode into rectangles' position
            for i in range(len(rectangles)):
                for e in range(width - rectangles[i][0] + 1):
                    if result[f"px{i + 1},{e}"] == False and result[f"px{i + 1},{e + 1}"] == True:
                        pos[i][0] = e + 1
                    if e == 0 and result[f"px{i + 1},{e}"] == True:
                        pos[i][0] = 0
                for f in range(height - rectangles[i][1] + 1):
                    if result[f"py{i + 1},{f}"] == False and result[f"py{i + 1},{f + 1}"] == True:
                        pos[i][1] = f + 1
                    if f == 0 and result[f"py{i + 1},{f}"] == True:
                        pos[i][1] = 0
                bins_used = []
            for j in range(max_bins):
                bins_used.append([i for i in range(len(rectangles)) if result[f"x{i + 1},{j + 1}"] == True])
            return(["sat", bins_used, pos, format(stop-start, '.6f'), len(variables), len(cnf.clauses)])

        else:
            timer.cancel()
            return("unsat")
        
    except:
        timer.cancel()
        return("timeout")
    
            
def BPP(W, H, items, n):
    items_area = [i[0] * i[1] for i in items]
    bin_area = W * H
    lower_bound = math.ceil(sum(items_area) / bin_area)
    for k in range(lower_bound, n + 1):
        print("Trying with", k, "bins")
        result = OPP(items, k, W, H)
        if result[0] == "sat":
            print("Solution found with", k, "bins")
            
            return result[1:]
        

def print_solution(bpp_result):
    if bpp_result is None:
        print("No solution found")
        return
    else:
        bins = bpp_result[0]
        pos = bpp_result[1]
        solver_time = bpp_result[2]
        num_variables = bpp_result[3]
        num_clauses = bpp_result[4]
        for i in range(len(bins)):
            print("Bin", i + 1, "contains items", [(j + 1) for j in bins[i]])
            for j in bins[i]:
                print("Item", j + 1, items[j], "at position", pos[j])
            display_solution((W, H), [items[j] for j in bins[i]], [pos[j] for j in bins[i]])
        print("--------------------")
        print("Solution found with", len(bins), "bins")
        print("Solver time:", solver_time)
        print("Number of variables:", num_variables)
        print("Number of clauses:", num_clauses)

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path = 'out/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)
    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")

def export_to_xlsx(bpp_result, filepath, time):
    result_dict = {}
    if bpp_result is None:
        result_dict = {
            "Type": "SAT no rotation ",
            "Dataset": filepath.split("/")[-1],
            "Number of items": n,
            "Minimize Bin": "-",  
            "Solver time": "timeout", 
            "Real time": "timeout",
            "Number of variables": "-", 
            "Number of clauses": "-"}
    else:
        bins, pos, solver_time, num_variables, num_clauses = bpp_result
        result_dict = {
            "Type": "SAT no rotation",
            "Dataset": filepath.split("/")[-1],
            "Number of items": n,
            "Minimize Bin": len(bins),  
            "Solver time": solver_time, 
            "Real time": time,
            "Number of variables": num_variables, 
            "Number of clauses": num_clauses}
    write_to_xlsx(result_dict)
n = 0 
def interrupt():
    print("Timeout")
    write_to_xlsx({
        "Type": "SAT no rotation",
        "Dataset": filepath.split("/")[-1],
        "Number of items": n,
        "Minimize Bin": "-",
        "Solver time": "timeout",
        "Real time": "timeout",
        "Number of variables": "-",
        "Number of clauses": "-"
    })
    os._exit(0)   
# Main
for i in range(1, 6):
    # timer = Timer(600, interrupt)
    # timer.start()
    filepath = f"input_data/class/CL_020_0{i}.txt"
    print(f"Processing file: {filepath}")
    input = read_file_instance(filepath)
    n = int(input[0])
    bin_size = input[1].split()
    W = int(bin_size[0])
    H = int(bin_size[1])
    items = [[int(val) for val in i.split()] for i in input[2:]]
    start = time.time()
    bpp_result = BPP(W, H, items, n)
    stop = time.time()
    export_to_xlsx(bpp_result, filepath, format(stop - start, '.6f'))
    print("Solver", format(stop - start, '.6f'))
    # print_solution(bpp_result)
