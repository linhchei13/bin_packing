from ortools.linear_solver import pywraplp
import sys
import pandas as pd
import os
import sys
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime 

def input_data(file_path):
    data = {}
    f = open(file_path,'r')
    
    data['size_item'] = []
    data['size_bin'] = []
    n = int(f.readline())
    W, H = map(int, f.readline().split())
    data['size_bin'].append([W, H])
    for i in range(1,n+1):
        line = f.readline().split()
        data['size_item'].append([int(line[0]),int(line[1])])
    
    return n,data,W,H

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path = 'out/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)
    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

    print(f"Result added to Excel file: {os.path.abspath(excel_file_path)}\n")
def main_solver(file_path, time_limit):
    n,data,W,H = input_data(file_path)
    k = n
    print('Number of items:',n)
    print('Size of bin:',W,H)
    solver = pywraplp.Solver.CreateSolver('SCIP')
    
    # Create variables
    M = 1000000

    x = {} # x[(i,m)] = 1 iff item i is packed in car m else 0
    # Ro represent for R in the presentation file/ pdf model file
    Ro = {} # if Ro = 1 then rotation = 90 degree, else 0
    l = {} # left coordination of item
    r = {} # right coordination of item
    t = {} # top coordination of item
    b = {} # bottom coodination of item

    for i in range(n):
        Ro[i] = solver.IntVar(0, 1, 'Ro[%i] '%i)

        # coordinate
        l[i] = solver.IntVar(0, W,'l[%i]' % i)
        r[i] = solver.IntVar(0, W,'r[%i]' % i)
        t[i] = solver.IntVar(0, H,'t[%i]' % i)
        b[i] = solver.IntVar(0, H,'b[%i]' % i)        

        solver.Add(r[i] == (1-Ro[i]) * data['size_item'][i][0] + Ro[i] * data['size_item'][i][1] + l[i])
        solver.Add(t[i] == (1-Ro[i]) * data['size_item'][i][1] + Ro[i] * data['size_item'][i][0] + b[i])

        for m in range(k):

            x[(i,m)] = solver.IntVar(0, 1, 'x_[%i]_[%i]' %(i,m))

            # item i must not exceed area of car
            solver.Add(r[i] <= (1-x[(i,m)]) * M + W)
            solver.Add(l[i] <= (1-x[(i,m)]) * M + W)
            solver.Add(t[i] <= (1-x[(i,m)]) * M + H)
            solver.Add(b[i] <= (1-x[(i,m)]) * M + H)    

    for i in range(n):
        solver.Add(sum(x[(i,m)] for m in range(k)) == 1)


    # if 2 items is packed in the same car, they must be not overlaped
    for i in range(n - 1):
        for j in range(i + 1, n):
            for m in range(k):
                e = solver.IntVar(0, 1, f'e[{i}][{j}]')
                solver.Add(e >= x[i,m] + x[j,m] - 1)
                solver.Add(e <= x[i,m])
                solver.Add(e <= x[j,m])

                # Binary variables for each constraint
                c1 = solver.IntVar(0, 1, f'c1[{i}][{j}]')
                c2 = solver.IntVar(0, 1, f'c2[{i}][{j}]')
                c3 = solver.IntVar(0, 1, f'c3[{i}][{j}]')
                c4 = solver.IntVar(0, 1, f'c4[{i}][{j}]')
                
                # Constraints that the binary variables must satisfy
                solver.Add(r[i] <= l[j] + M * (1 - c1))
                solver.Add(r[j] <= l[i] + M * (1 - c2))
                solver.Add(t[i] <= b[j] + M * (1 - c3))
                solver.Add(t[j] <= b[i] + M * (1 - c4))

                solver.Add(c1 + c2 + c3 + c4 + (1-e)*M >= 1 )
                solver.Add(c1 + c2 + c3 + c4 <= e*M )

    # find cars be used
    z = {} # z[m] = 1 iff car m be used
    for m in range(k):
        z[m] = solver.IntVar(0, 1, 'z[%i] ' %m)
        # if sum(x[i][m]) >= 1 then car m be used => z[m] = 1
        # else, z[m] = 0

        q = solver.IntVar(0,n,f'q[{m}]')
        solver.Add(q == sum(x[(i,m)] for i in range(n)))
        # car m be used iff there are at least 1 item be packed in car m, so sum(x[(i,m)] for i in range(n)) != 0 
        
        # q = 0 => z[m] = 0
        # q != 0 => z[m] = 1
        solver.Add(z[m] <= q * M)
        solver.Add(q <= z[m] * M)

    # objective
    bin = sum(z[m] for m in range(k))
    solver.Minimize(bin)
    solver.set_time_limit(time_limit * 1000)

    status = solver.Solve()
    print(status)
    result_dict = {}
    if solver.Solve() == pywraplp.Solver.OPTIMAL or solver.Solve() == pywraplp.Solver.FEASIBLE:
        print('--------------Solution Found--------------')
        for i in range(n):
            print(f'put item {i+1} with rotation {int(Ro[i].solution_value())}', end=' ') 
            for j in range(k):
                if x[i,j].solution_value() ==1:
                    print(f'in bin {j+1}', end=' ')
            print(f'at ({l[i].solution_value()},{b[i].solution_value()})')
        print(f'Number of bin used  :',int(sum(z[m].solution_value() for m in range(k))))
        print('----------------Statistics----------------')
        if status == pywraplp.Solver.OPTIMAL:
            print('Status              : OPTIMAL')
        else:
            print('Status              : FEASIBLE')
        print(f'Time limit          : {time_limit}')
        print(f'Running time        : {solver.WallTime() / 1000}')
        return n, int(sum(z[m].solution_value() for m in range(k))), format(solver.WallTime() / 1000, '.6f')
    else:
        print('NO SOLUTIONS')
        return n, '-', format(solver.WallTime() / 1000, '.6f')
    
if __name__ == '__main__':
    try:
        # Get input file path
        time_limit = int(sys.argv[2])
    except IndexError:
        time_limit = 600
       
    # Create solver
    
    file_path = f'input_data/BENG/BENG05.txt'
    print("Reading file: ", file_path.split("/")[-1])
    start = time.time()
    n, n_bins, solver_time = main_solver(file_path, time_limit)
    stop = time.time()
    result_dict = {
        "Type": "Ortools MIP",
        "Data": file_path.split("/")[-1],
        "Number of items": n,
        "Bins": n_bins,  
        "Solver time": solver_time,
        "Real time": format(stop - start, '.6f')
    }
    write_to_xlsx(result_dict)
    



